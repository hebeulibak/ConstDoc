from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
from functools import wraps
import os
import uuid
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'stroi dock-super-secret-key-2024'

# Конфигурация БД
basedir = os.path.abspath(os.path.dirname(__file__))

# Создаём папку instance если её нет
instance_path = os.path.join(basedir, 'instance')
if not os.path.exists(instance_path):
    os.makedirs(instance_path)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(instance_path, 'stroi_dock.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Конфигурация для загрузки файлов
UPLOAD_FOLDER = os.path.join(basedir, 'uploads')
DOCUMENTS_FOLDER = os.path.join(UPLOAD_FOLDER, 'documents')
PHOTOS_FOLDER = os.path.join(UPLOAD_FOLDER, 'photos')
MEDIA_FOLDER = os.path.join(UPLOAD_FOLDER, 'media')
AVATAR_FOLDER = os.path.join(UPLOAD_FOLDER, 'avatars')

for folder in [UPLOAD_FOLDER, DOCUMENTS_FOLDER, PHOTOS_FOLDER, MEDIA_FOLDER, AVATAR_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB для видео

ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'gif', 'mp4', 'webm', 'avi', 'mov', 'mkv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)

# ============ ДЕКОРАТОР АВТОРИЗАЦИИ ============
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Пожалуйста, авторизуйтесь', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Пожалуйста, авторизуйтесь', 'warning')
            return redirect(url_for('login'))
        if session.get('user_role') != 'admin':
            flash('Доступ запрещён. Требуются права администратора', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ============ МОДЕЛИ ============

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), default='user')
    created_at = db.Column(db.DateTime, default=datetime.now)

class Site(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(300))
    customer = db.Column(db.String(200))
    contractor = db.Column(db.String(200))
    start_date = db.Column(db.String(50))
    end_date = db.Column(db.String(50))
    status = db.Column(db.String(50), default='active')
    created_at = db.Column(db.DateTime, default=datetime.now)
    avatar_filename = db.Column(db.String(300))
    
    entries = db.relationship('JournalEntry', backref='site', lazy=True, cascade='all, delete-orphan')
    documents = db.relationship('Document', backref='site', lazy=True, cascade='all, delete-orphan')
    photos = db.relationship('Photo', backref='site', lazy=True, cascade='all, delete-orphan')
    media = db.relationship('Media', backref='site', lazy=True, cascade='all, delete-orphan')

class JournalEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    date = db.Column(db.Date, nullable=False, default=datetime.now)
    work_type = db.Column(db.String(200), nullable=False)
    work_description = db.Column(db.Text)
    location = db.Column(db.String(200))
    executor = db.Column(db.String(200))
    responsible_person = db.Column(db.String(200))
    workers_count = db.Column(db.Integer, default=1)
    start_time = db.Column(db.String(10))
    end_time = db.Column(db.String(10))
    volume = db.Column(db.Float, default=0)
    volume_unit = db.Column(db.String(20), default='м³')
    equipment_used = db.Column(db.String(300))
    materials_used = db.Column(db.String(300))
    weather = db.Column(db.String(100))
    temperature = db.Column(db.String(50))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.strftime('%Y-%m-%d'),
            'work_type': self.work_type,
            'work_description': self.work_description or '',
            'location': self.location or '',
            'executor': self.executor or '',
            'responsible_person': self.responsible_person or '',
            'workers_count': self.workers_count,
            'start_time': self.start_time or '',
            'end_time': self.end_time or '',
            'volume': self.volume,
            'volume_unit': self.volume_unit,
            'equipment_used': self.equipment_used or '',
            'materials_used': self.materials_used or '',
            'weather': self.weather or '',
            'temperature': self.temperature or '',
            'notes': self.notes or ''
        }

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    doc_type = db.Column(db.String(50))  # incoming, outgoing, executive
    title = db.Column(db.String(300), nullable=False)
    description = db.Column(db.Text)
    filename = db.Column(db.String(300))
    file_path = db.Column(db.String(500))
    file_size = db.Column(db.Integer)
    uploaded_by = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)

class Photo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    title = db.Column(db.String(300))
    description = db.Column(db.Text)
    filename = db.Column(db.String(300))
    file_path = db.Column(db.String(500))
    latitude = db.Column(db.Float)
    longitude = db.Column(db.Float)
    location_name = db.Column(db.String(300))
    photo_date = db.Column(db.DateTime, default=datetime.now)
    uploaded_by = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)

class Media(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    media_type = db.Column(db.String(20))  # photo или video
    title = db.Column(db.String(300))
    description = db.Column(db.Text)
    filename = db.Column(db.String(300))
    file_path = db.Column(db.String(500))
    latitude = db.Column(db.Float)
    longitude = db.Column(db.Float)
    location_name = db.Column(db.String(300))
    upload_date = db.Column(db.DateTime, default=datetime.now)
    uploaded_by = db.Column(db.String(100))

class Normative(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), nullable=False)
    title = db.Column(db.String(300), nullable=False)
    description = db.Column(db.Text)
    doc_type = db.Column(db.String(50))
    category = db.Column(db.String(100))
    status = db.Column(db.String(50), default='действует')
    actual_date = db.Column(db.String(50))
    tags = db.Column(db.String(500))
    file_url = db.Column(db.String(500))
    is_favorite = db.Column(db.Boolean, default=False)
    views = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)

class KnowledgeArticle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(300), nullable=False)
    content = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(100))
    author = db.Column(db.String(100))
    views = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)


# ============ ГРАФИК РАБОТ ============

class ScheduleTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_id = db.Column(db.Integer, db.ForeignKey('site.id'), nullable=False)
    
    # Основные поля
    order_num = db.Column(db.Integer, default=0)
    task_name = db.Column(db.String(300), nullable=False)
    description = db.Column(db.Text)
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    duration = db.Column(db.Integer)
    
    # Зависимости (ID работ-предшественников через запятую)
    dependencies = db.Column(db.String(500))
    
    # Статус и прогресс
    status = db.Column(db.String(50), default='plan')
    progress = db.Column(db.Integer, default=0)
    actual_start = db.Column(db.Date)
    actual_end = db.Column(db.Date)
    
    # Ресурсы
    required_materials = db.Column(db.String(500))
    responsible = db.Column(db.String(200))
    notes = db.Column(db.Text)
    
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    site = db.relationship('Site', backref='schedule_tasks', lazy=True)

    def to_dict(self):
        return {
            'id': self.id,
            'order_num': self.order_num,
            'task_name': self.task_name,
            'description': self.description,
            'start_date': self.start_date.strftime('%Y-%m-%d') if self.start_date else '',
            'end_date': self.end_date.strftime('%Y-%m-%d') if self.end_date else '',
            'duration': self.duration,
            'dependencies': self.dependencies,
            'status': self.status,
            'progress': self.progress,
            'actual_start': self.actual_start.strftime('%Y-%m-%d') if self.actual_start else '',
            'actual_end': self.actual_end.strftime('%Y-%m-%d') if self.actual_end else '',
            'required_materials': self.required_materials,
            'responsible': self.responsible,
            'notes': self.notes
        }


# ============ СОЗДАНИЕ ТАБЛИЦ И ТЕСТОВЫХ ДАННЫХ ============

with app.app_context():
    db.create_all()
    
    # Создание тестового пользователя
    if User.query.count() == 0:
        test_user = User(username='admin', password='admin123', role='admin')
        db.session.add(test_user)
        db.session.commit()
        print("✅ Тестовый пользователь создан: admin / admin123")
    
    # Создание тестовых объектов
    if Site.query.count() == 0:
        test_sites = [
            Site(
                name='Жилой комплекс "Янтарный"',
                address='г. Москва, ул. Строителей, 15',
                customer='ООО "Девелопмент Групп"',
                contractor='АО "СтройИнвест"',
                start_date=datetime.now().strftime('%Y-%m-%d'),
                status='active'
            ),
            Site(
                name='Бизнес-центр "Плаза"',
                address='г. Санкт-Петербург, Невский пр., 100',
                customer='ЗАО "Коммерц Недвижимость"',
                contractor='ООО "СеверСтрой"',
                start_date='2024-01-15',
                status='active'
            ),
            Site(
                name='Торговый центр "Гранд"',
                address='г. Казань, ул. Ямашева, 50',
                customer='ООО "ТоргИнвест"',
                contractor='ПАО "ТатарСтрой"',
                start_date='2023-06-01',
                end_date='2024-03-15',
                status='completed'
            )
        ]
        for site in test_sites:
            db.session.add(site)
        db.session.commit()
        print("✅ Тестовые объекты созданы")
    
    # Создание тестовых нормативов
    if Normative.query.count() == 0:
        normatives_data = [
            Normative(
                code="СП 48.13330.2019",
                title="Организация строительства",
                description="Актуализированная редакция СНиП 12-01-2004. Основные требования к организации строительного производства.",
                doc_type="СП",
                category="Строительство",
                status="действует",
                actual_date="2025",
                tags="организация строительства, ПОС, ППР",
                is_favorite=True
            ),
            Normative(
                code="ГОСТ Р 21.1101-2013",
                title="СПДС. Основные требования к проектной документации",
                description="Требования к составу и оформлению проектной документации.",
                doc_type="ГОСТ",
                category="Проектирование",
                status="действует",
                actual_date="2025",
                tags="проектная документация, СПДС",
                is_favorite=True
            ),
            Normative(
                code="ФЗ-384",
                title="Технический регламент о безопасности зданий и сооружений",
                description="Федеральный закон о требованиях безопасности зданий и сооружений.",
                doc_type="Закон",
                category="Законодательство",
                status="действует",
                actual_date="2025",
                tags="технический регламент, безопасность",
                is_favorite=True
            )
        ]
        for n in normatives_data:
            db.session.add(n)
        db.session.commit()
        print(f"✅ Тестовые нормативы созданы: {len(normatives_data)} документов")
    
    # Создание тестовых статей
    if KnowledgeArticle.query.count() == 0:
        articles = [
            KnowledgeArticle(
                title='Как правильно вести журнал работ',
                content='Журнал работ должен содержать ежедневные записи о выполненных работах, используемых материалах и оборудовании.',
                category='Практика',
                author='Главный инженер'
            ),
            KnowledgeArticle(
                title='Требования к технике безопасности на стройплощадке',
                content='Все работники должны быть обеспечены средствами индивидуальной защиты.',
                category='Безопасность',
                author='Инженер по охране труда'
            )
        ]
        for a in articles:
            db.session.add(a)
        db.session.commit()
        print(f"✅ Тестовые статьи созданы: {len(articles)} статей")

# ============ КАЛЬКУЛЯТОР МАТЕРИАЛОВ ============
import json
import math

# ============ ОБЩЕСТРОИТЕЛЬНЫЕ РАБОТЫ ============
CONSTRUCTION_MATERIALS = {
    # Земляные работы
    "excavation": {
        "name": "Разработка котлована",
        "category": "Земляные работы",
        "unit": "м³",
        "main_material": {"name": "выемка грунта", "per_unit": 1.0, "unit": "м³"},
        "additional": None,
        "waste_factor": 1.0
    },
    "backfill": {
        "name": "Обратная засыпка пазух",
        "category": "Земляные работы",
        "unit": "м³",
        "main_material": {"name": "песок/грунт", "per_unit": 1.0, "unit": "м³"},
        "additional": None,
        "waste_factor": 1.05
    },
    
    # Бетонные работы
    "concrete_foundation": {
        "name": "Заливка фундамента (бетон М300)",
        "category": "Бетонные работы",
        "unit": "м³",
        "main_material": {"name": "бетон", "per_unit": 1.0, "unit": "м³"},
        "additional": [{"name": "арматура", "per_unit": 80, "unit": "кг"}],
        "waste_factor": 1.03
    },
    "concrete_floor": {
        "name": "Монолитное перекрытие",
        "category": "Бетонные работы",
        "unit": "м³",
        "main_material": {"name": "бетон", "per_unit": 1.0, "unit": "м³"},
        "additional": [
            {"name": "арматура", "per_unit": 100, "unit": "кг"},
            {"name": "опалубка", "per_unit": 3.5, "unit": "м²"}
        ],
        "waste_factor": 1.03
    },
    
    # Каменные работы
    "brick_1": {
        "name": "Кладка стен из кирпича (1 кирпич, 250мм)",
        "category": "Каменные работы",
        "unit": "м²",
        "main_material": {"name": "кирпич", "per_unit": 102, "unit": "шт"},
        "additional": [{"name": "раствор", "per_unit": 0.1, "unit": "м³"}],
        "waste_factor": 1.05
    },
    "brick_half": {
        "name": "Кладка перегородок (0.5 кирпича, 120мм)",
        "category": "Каменные работы",
        "unit": "м²",
        "main_material": {"name": "кирпич", "per_unit": 51, "unit": "шт"},
        "additional": [{"name": "раствор", "per_unit": 0.05, "unit": "м³"}],
        "waste_factor": 1.05
    },
    "block_wall": {
        "name": "Кладка стен из газобетонных блоков",
        "category": "Каменные работы",
        "unit": "м³",
        "main_material": {"name": "газоблоки", "per_unit": 28, "unit": "шт"},
        "additional": [{"name": "клей для блоков", "per_unit": 25, "unit": "кг"}],
        "waste_factor": 1.05
    },
    
    # Отделочные работы
    "plaster": {
        "name": "Штукатурка стен (толщина 2см)",
        "category": "Отделочные работы",
        "unit": "м²",
        "main_material": {"name": "сухая штукатурная смесь", "per_unit": 15, "unit": "кг"},
        "additional": [{"name": "грунтовка", "per_unit": 0.2, "unit": "л"}],
        "waste_factor": 1.05
    },
    "paint": {
        "name": "Покраска стен (2 слоя)",
        "category": "Отделочные работы",
        "unit": "м²",
        "main_material": {"name": "краска", "per_unit": 0.2, "unit": "л"},
        "additional": None,
        "waste_factor": 1.03
    },
    "wallpaper": {
        "name": "Оклейка обоями",
        "category": "Отделочные работы",
        "unit": "м²",
        "main_material": {"name": "обои", "per_unit": 1.1, "unit": "м²"},
        "additional": [{"name": "клей для обоев", "per_unit": 0.1, "unit": "кг"}],
        "waste_factor": 1.05
    },
    "tile_floor": {
        "name": "Укладка плитки на пол",
        "category": "Отделочные работы",
        "unit": "м²",
        "main_material": {"name": "плитка керамическая", "per_unit": 1.05, "unit": "м²"},
        "additional": [
            {"name": "плиточный клей", "per_unit": 4.5, "unit": "кг"},
            {"name": "затирка", "per_unit": 0.3, "unit": "кг"}
        ],
        "waste_factor": 1.07
    },
    "screed": {
        "name": "Стяжка пола (пескобетон, толщина 5см)",
        "category": "Отделочные работы",
        "unit": "м²",
        "main_material": {"name": "пескобетон", "per_unit": 90, "unit": "кг"},  # 1800 кг/м³ * 0.05м
        "additional": None,
        "waste_factor": 1.05
    },
    
    # Кровельные работы
    "roof_metal": {
        "name": "Монтаж металлочерепицы",
        "category": "Кровельные работы",
        "unit": "м²",
        "main_material": {"name": "металлочерепица", "per_unit": 1.05, "unit": "м²"},
        "additional": [
            {"name": "саморезы", "per_unit": 8, "unit": "шт"},
            {"name": "гидроизоляция", "per_unit": 1.05, "unit": "м²"}
        ],
        "waste_factor": 1.05
    },
    "roof_soft": {
        "name": "Монтаж мягкой кровли",
        "category": "Кровельные работы",
        "unit": "м²",
        "main_material": {"name": "мягкая черепица", "per_unit": 1.05, "unit": "м²"},
        "additional": [{"name": "подкладочный ковер", "per_unit": 1.05, "unit": "м²"}],
        "waste_factor": 1.05
    },
}

# ============ МЕТАЛЛОПРОКАТ (ДЛЯ СВАРОЧНЫХ РАБОТ) ============
METAL_SECTIONS = {
    # Уголки
    "angle_25x3": {"name": "Уголок 25x25x3 мм", "weight_per_m": 1.12, "unit": "кг/м"},
    "angle_30x3": {"name": "Уголок 30x30x3 мм", "weight_per_m": 1.36, "unit": "кг/м"},
    "angle_40x4": {"name": "Уголок 40x40x4 мм", "weight_per_m": 2.42, "unit": "кг/м"},
    "angle_50x5": {"name": "Уголок 50x50x5 мм", "weight_per_m": 3.77, "unit": "кг/м"},
    "angle_63x6": {"name": "Уголок 63x63x6 мм", "weight_per_m": 5.72, "unit": "кг/м"},
    "angle_75x7": {"name": "Уголок 75x75x7 мм", "weight_per_m": 7.94, "unit": "кг/м"},
    "angle_90x8": {"name": "Уголок 90x90x8 мм", "weight_per_m": 10.8, "unit": "кг/м"},
    
    # Профильные трубы (квадрат)
    "square_tube_40x40x2": {"name": "Труба профильная 40x40x2 мм", "weight_per_m": 2.37, "unit": "кг/м"},
    "square_tube_50x50x2": {"name": "Труба профильная 50x50x2 мм", "weight_per_m": 2.98, "unit": "кг/м"},
    "square_tube_60x60x3": {"name": "Труба профильная 60x60x3 мм", "weight_per_m": 5.25, "unit": "кг/м"},
    "square_tube_80x80x3": {"name": "Труба профильная 80x80x3 мм", "weight_per_m": 7.13, "unit": "кг/м"},
    "square_tube_100x100x4": {"name": "Труба профильная 100x100x4 мм", "weight_per_m": 11.8, "unit": "кг/м"},
    
    # Профильные трубы (прямоугольник)
    "rect_tube_60x40x3": {"name": "Труба профильная 60x40x3 мм", "weight_per_m": 4.36, "unit": "кг/м"},
    "rect_tube_80x40x3": {"name": "Труба профильная 80x40x3 мм", "weight_per_m": 5.31, "unit": "кг/м"},
    "rect_tube_100x50x4": {"name": "Труба профильная 100x50x4 мм", "weight_per_m": 8.69, "unit": "кг/м"},
    "rect_tube_120x60x4": {"name": "Труба профильная 120x60x4 мм", "weight_per_m": 10.6, "unit": "кг/м"},
    
    # Швеллеры
    "channel_8p": {"name": "Швеллер 8П", "weight_per_m": 7.05, "unit": "кг/м"},
    "channel_10p": {"name": "Швеллер 10П", "weight_per_m": 8.59, "unit": "кг/м"},
    "channel_12p": {"name": "Швеллер 12П", "weight_per_m": 10.4, "unit": "кг/м"},
    "channel_14p": {"name": "Швеллер 14П", "weight_per_m": 12.3, "unit": "кг/м"},
    "channel_16p": {"name": "Швеллер 16П", "weight_per_m": 14.2, "unit": "кг/м"},
    "channel_18p": {"name": "Швеллер 18П", "weight_per_m": 16.3, "unit": "кг/м"},
    "channel_20p": {"name": "Швеллер 20П", "weight_per_m": 18.4, "unit": "кг/м"},
    
    # Двутавры
    "beam_10": {"name": "Двутавр 10", "weight_per_m": 9.46, "unit": "кг/м"},
    "beam_12": {"name": "Двутавр 12", "weight_per_m": 11.5, "unit": "кг/м"},
    "beam_14": {"name": "Двутавр 14", "weight_per_m": 13.7, "unit": "кг/м"},
    "beam_16": {"name": "Двутавр 16", "weight_per_m": 15.9, "unit": "кг/м"},
    "beam_18": {"name": "Двутавр 18", "weight_per_m": 18.4, "unit": "кг/м"},
    "beam_20": {"name": "Двутавр 20", "weight_per_m": 21.0, "unit": "кг/м"},
    "beam_24": {"name": "Двутавр 24", "weight_per_m": 27.0, "unit": "кг/м"},
    "beam_30": {"name": "Двутавр 30", "weight_per_m": 36.5, "unit": "кг/м"},
    
    # Круглые трубы
    "round_tube_32x2": {"name": "Труба круглая 32x2 мм", "weight_per_m": 1.48, "unit": "кг/м"},
    "round_tube_42x2.5": {"name": "Труба круглая 42x2.5 мм", "weight_per_m": 2.44, "unit": "кг/м"},
    "round_tube_57x3": {"name": "Труба круглая 57x3 мм", "weight_per_m": 3.99, "unit": "кг/м"},
    "round_tube_76x3": {"name": "Труба круглая 76x3 мм", "weight_per_m": 5.40, "unit": "кг/м"},
    "round_tube_89x4": {"name": "Труба круглая 89x4 мм", "weight_per_m": 8.38, "unit": "кг/м"},
    "round_tube_108x4": {"name": "Труба круглая 108x4 мм", "weight_per_m": 10.2, "unit": "кг/м"},
    "round_tube_133x5": {"name": "Труба круглая 133x5 мм", "weight_per_m": 15.8, "unit": "кг/м"},
    
    # Арматура
    "rebar_6": {"name": "Арматура А500С Ø6 мм", "weight_per_m": 0.222, "unit": "кг/м"},
    "rebar_8": {"name": "Арматура А500С Ø8 мм", "weight_per_m": 0.395, "unit": "кг/м"},
    "rebar_10": {"name": "Арматура А500С Ø10 мм", "weight_per_m": 0.617, "unit": "кг/м"},
    "rebar_12": {"name": "Арматура А500С Ø12 мм", "weight_per_m": 0.888, "unit": "кг/м"},
    "rebar_14": {"name": "Арматура А500С Ø14 мм", "weight_per_m": 1.21, "unit": "кг/м"},
    "rebar_16": {"name": "Арматура А500С Ø16 мм", "weight_per_m": 1.58, "unit": "кг/м"},
    "rebar_18": {"name": "Арматура А500С Ø18 мм", "weight_per_m": 2.00, "unit": "кг/м"},
    "rebar_20": {"name": "Арматура А500С Ø20 мм", "weight_per_m": 2.47, "unit": "кг/м"},
    "rebar_22": {"name": "Арматура А500С Ø22 мм", "weight_per_m": 2.98, "unit": "кг/м"},
    "rebar_25": {"name": "Арматура А500С Ø25 мм", "weight_per_m": 3.85, "unit": "кг/м"},
    "rebar_28": {"name": "Арматура А500С Ø28 мм", "weight_per_m": 4.83, "unit": "кг/м"},
    "rebar_32": {"name": "Арматура А500С Ø32 мм", "weight_per_m": 6.31, "unit": "кг/м"},
    
    # Листовой металл (специальный тип)
    "sheet_metal": {
        "name": "Листовой металл",
        "type": "sheet",
        "density": 7850,  # кг/м³
        "unit_input": "м²"
    }
}

# ============ АСФАЛЬТИРОВАНИЕ ============
ASPHALT_TYPES = {
    "asphalt_fine": {
        "name": "Мелкозернистый асфальт (тип Б)",
        "consumption_per_cm": 24.5,  # кг/м² на 1 см толщины
        "unit": "кг/м²"
    },
    "asphalt_coarse": {
        "name": "Крупнозернистый асфальт",
        "consumption_per_cm": 24.2,
        "unit": "кг/м²"
    },
    "asphalt_shma": {
        "name": "Щебеночно-мастичная смесь (ЩМА-15)",
        "consumption_per_cm": 25.8,
        "unit": "кг/м²"
    }
}

# Объединяем всё в один словарь для удобного доступа
MATERIALS_CONFIG = {**CONSTRUCTION_MATERIALS, **METAL_SECTIONS, **ASPHALT_TYPES}

# ============ РОУТЫ АВТОРИЗАЦИИ ============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            session['logged_in'] = True
            session['username'] = username
            session['user_role'] = user.role
            flash(f'Добро пожаловать, {username}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверное имя пользователя или пароль', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('login'))

# ============ РОУТЫ ОСНОВНЫЕ ============

@app.route('/')
@login_required
def index():
    sites = Site.query.all()
    
    active_count = Site.query.filter_by(status='active').count()
    completed_count = Site.query.filter_by(status='completed').count()
    paused_count = Site.query.filter_by(status='paused').count()
    
    current_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('index.html', 
                         sites=sites, 
                         active_count=active_count,
                         completed_count=completed_count,
                         paused_count=paused_count,
                         current_date=current_date)

@app.route('/add-site', methods=['POST'])
@login_required
def add_site_route():
    data = request.form
    site = Site(
        name=data['name'],
        address=data.get('address', ''),
        customer=data.get('customer', ''),
        contractor=data.get('contractor', ''),
        start_date=data.get('start_date', ''),
        end_date=data.get('end_date', ''),
        status=data.get('status', 'active')
    )
    db.session.add(site)
    db.session.commit()
    flash(f'✅ Объект "{site.name}" успешно добавлен', 'success')
    return redirect(url_for('index'))

@app.route('/site/delete/<int:site_id>')
@login_required
def delete_site(site_id):
    site = Site.query.get_or_404(site_id)
    name = site.name
    db.session.delete(site)
    db.session.commit()
    flash(f'🗑️ Объект "{name}" удалён', 'danger')
    return redirect(url_for('index'))

# ============ СТРАНИЦА ОБЪЕКТА ============

@app.route('/site/<int:site_id>')
@login_required
def site_detail(site_id):
    site = Site.query.get_or_404(site_id)
    
    stats = {
        'journal_count': JournalEntry.query.filter_by(site_id=site_id).count(),
        'incoming_count': Document.query.filter_by(site_id=site_id, doc_type='incoming').count(),
        'outgoing_count': Document.query.filter_by(site_id=site_id, doc_type='outgoing').count(),
        'executive_count': Document.query.filter_by(site_id=site_id, doc_type='executive').count()
    }
    
# Получаем сохранённые расчёты из сессии для отображения счётчика
    saved_calculations = session.get(f'calculations_{site_id}', [])
    stats['calculator_count'] = len(saved_calculations)

    return render_template('site_detail.html', site=site, stats=stats)

@app.route('/journal/embed/<int:site_id>')
@login_required
def journal_embed(site_id):
    site = Site.query.get_or_404(site_id)
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    work_type = request.args.get('work_type')
    responsible = request.args.get('responsible')
    
    query = JournalEntry.query.filter_by(site_id=site_id)
    
    if date_from:
        query = query.filter(JournalEntry.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(JournalEntry.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if work_type:
        query = query.filter_by(work_type=work_type)
    if responsible:
        query = query.filter_by(responsible_person=responsible)
    
    page = request.args.get('page', 1, type=int)
    per_page = 15
    pagination = query.order_by(JournalEntry.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    entries = pagination.items
    
    total_entries = query.count()
    week_ago = datetime.now().date() - timedelta(days=7)
    weekly_count = JournalEntry.query.filter(
        JournalEntry.site_id == site_id,
        JournalEntry.date >= week_ago
    ).count()
    
    all_work_types = [wt[0] for wt in db.session.query(JournalEntry.work_type).filter_by(site_id=site_id).distinct()]
    responsible_persons = [r[0] for r in db.session.query(JournalEntry.responsible_person).filter_by(site_id=site_id).distinct() if r[0]]
    
    request_args = request.args.to_dict()
    if 'page' in request_args:
        del request_args['page']
    
    return render_template('journal_embed.html',
                         site=site,
                         entries=entries,
                         pagination=pagination,
                         stats={'total_entries': total_entries, 'weekly_count': weekly_count},
                         all_work_types=all_work_types,
                         responsible_persons=responsible_persons,
                         date_from=date_from or '',
                         date_to=date_to or '',
                         selected_work_type=work_type or '',
                         selected_responsible=responsible or '',
                         request_args=request_args)

# ============ API ДЛЯ ДОКУМЕНТОВ ============

@app.route('/api/documents/<int:site_id>')
@login_required
def api_documents(site_id):
    doc_type = request.args.get('type')
    query = Document.query.filter_by(site_id=site_id)
    if doc_type:
        query = query.filter_by(doc_type=doc_type)
    
    documents = query.order_by(Document.created_at.desc()).all()
    return jsonify([{
        'id': d.id,
        'title': d.title,
        'description': d.description,
        'filename': d.filename,
        'uploaded_by': d.uploaded_by,
        'created_at': d.created_at.strftime('%d.%m.%Y %H:%M')
    } for d in documents])

@app.route('/api/photos/<int:site_id>')
@login_required
def api_photos(site_id):
    photos = Photo.query.filter_by(site_id=site_id).order_by(Photo.photo_date.desc()).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'description': p.description,
        'filename': p.filename,
        'latitude': p.latitude,
        'longitude': p.longitude,
        'location_name': p.location_name,
        'photo_date': p.photo_date.strftime('%d.%m.%Y %H:%M'),
        'uploaded_by': p.uploaded_by
    } for p in photos])

@app.route('/upload-document', methods=['POST'])
@login_required
def upload_document():
    site_id = request.form.get('site_id')
    doc_type = request.form.get('doc_type')
    title = request.form.get('title')
    description = request.form.get('description', '')
    
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не выбран'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Неподдерживаемый формат файла'}), 400
    
    ext = file.filename.rsplit('.', 1)[1].lower()
    new_filename = f"{uuid.uuid4().hex}.{ext}"
    
    type_folder = os.path.join(DOCUMENTS_FOLDER, doc_type)
    if not os.path.exists(type_folder):
        os.makedirs(type_folder)
    
    file_path = os.path.join(type_folder, new_filename)
    file.save(file_path)
    
    doc = Document(
        site_id=site_id,
        doc_type=doc_type,
        title=title,
        description=description,
        filename=new_filename,
        file_path=file_path,
        file_size=os.path.getsize(file_path),
        uploaded_by=session.get('username', 'Unknown')
    )
    db.session.add(doc)
    db.session.commit()
    
    return jsonify({'success': True, 'id': doc.id})

@app.route('/upload-photo', methods=['POST'])
@login_required
def upload_photo():
    site_id = request.form.get('site_id')
    title = request.form.get('title')
    description = request.form.get('description', '')
    latitude = request.form.get('latitude')
    longitude = request.form.get('longitude')
    
    if 'photo' not in request.files:
        return jsonify({'error': 'Фото не выбрано'}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'error': 'Фото не выбрано'}), 400
    
    ext = file.filename.rsplit('.', 1)[1].lower()
    new_filename = f"{uuid.uuid4().hex}.{ext}"
    
    site_photos_folder = os.path.join(PHOTOS_FOLDER, str(site_id))
    if not os.path.exists(site_photos_folder):
        os.makedirs(site_photos_folder)
    
    file_path = os.path.join(site_photos_folder, new_filename)
    file.save(file_path)
    
    photo = Photo(
        site_id=site_id,
        title=title,
        description=description,
        filename=new_filename,
        file_path=file_path,
        latitude=float(latitude) if latitude else None,
        longitude=float(longitude) if longitude else None,
        uploaded_by=session.get('username', 'Unknown')
    )
    db.session.add(photo)
    db.session.commit()
    
    return jsonify({'success': True, 'id': photo.id})

@app.route('/delete-document/<int:doc_id>', methods=['POST'])
@login_required
def delete_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    if os.path.exists(doc.file_path):
        os.remove(doc.file_path)
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/delete-photo/<int:photo_id>', methods=['POST'])
@login_required
def delete_photo(photo_id):
    photo = Photo.query.get_or_404(photo_id)
    if os.path.exists(photo.file_path):
        os.remove(photo.file_path)
    db.session.delete(photo)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/download/<int:doc_id>')
@login_required
def download_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    return send_file(doc.file_path, as_attachment=True, download_name=doc.title)

# ============ АВАТАР ОБЪЕКТА ============

@app.route('/upload-avatar/<int:site_id>', methods=['POST'])
@login_required
def upload_avatar(site_id):
    site = Site.query.get_or_404(site_id)
    if 'avatar' not in request.files:
        flash('Файл не выбран', 'danger')
        return redirect(url_for('site_detail', site_id=site_id))
    
    file = request.files['avatar']
    if file.filename == '':
        flash('Файл не выбран', 'danger')
        return redirect(url_for('site_detail', site_id=site_id))
    
    if site.avatar_filename:
        old_path = os.path.join(AVATAR_FOLDER, site.avatar_filename)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    ext = file.filename.rsplit('.', 1)[1].lower()
    new_filename = f"avatar_{site_id}_{uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(AVATAR_FOLDER, new_filename)
    file.save(file_path)
    
    site.avatar_filename = new_filename
    db.session.commit()
    flash('✅ Аватар объекта обновлён', 'success')
    return redirect(url_for('site_detail', site_id=site_id))

# ============ МЕДИАФАЙЛЫ (ФОТО/ВИДЕО) ============

@app.route('/api/media/<int:site_id>')
@login_required
def api_media(site_id):
    media = Media.query.filter_by(site_id=site_id).order_by(Media.upload_date.desc()).all()
    return jsonify([{
        'id': m.id,
        'media_type': m.media_type,
        'title': m.title,
        'description': m.description,
        'filename': m.filename,
        'location_name': m.location_name,
        'upload_date': m.upload_date.strftime('%d.%m.%Y %H:%M')
    } for m in media])

@app.route('/upload-media', methods=['POST'])
@login_required
def upload_media():
    site_id = request.form.get('site_id')
    media_type = request.form.get('media_type')
    title = request.form.get('title')
    description = request.form.get('description', '')
    latitude = request.form.get('latitude')
    longitude = request.form.get('longitude')
    
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не выбран'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    
    allowed_video = {'mp4', 'webm', 'avi', 'mov', 'mkv'}
    allowed_photo = {'jpg', 'jpeg', 'png', 'gif'}
    
    ext = file.filename.rsplit('.', 1)[1].lower()
    
    if media_type == 'photo' and ext not in allowed_photo:
        return jsonify({'error': 'Неподдерживаемый формат фото'}), 400
    elif media_type == 'video' and ext not in allowed_video:
        return jsonify({'error': 'Неподдерживаемый формат видео'}), 400
    
    new_filename = f"{uuid.uuid4().hex}.{ext}"
    
    media_folder = os.path.join(MEDIA_FOLDER, str(site_id))
    if not os.path.exists(media_folder):
        os.makedirs(media_folder)
    
    file_path = os.path.join(media_folder, new_filename)
    file.save(file_path)
    
    media = Media(
        site_id=site_id,
        media_type=media_type,
        title=title,
        description=description,
        filename=new_filename,
        file_path=file_path,
        latitude=float(latitude) if latitude else None,
        longitude=float(longitude) if longitude else None,
        uploaded_by=session.get('username', 'Unknown')
    )
    db.session.add(media)
    db.session.commit()
    
    return jsonify({'success': True, 'id': media.id})

@app.route('/delete-media/<int:media_id>', methods=['POST'])
@login_required
def delete_media(media_id):
    media = Media.query.get_or_404(media_id)
    if os.path.exists(media.file_path):
        os.remove(media.file_path)
    db.session.delete(media)
    db.session.commit()
    return jsonify({'success': True})

# ============ ЭКСПОРТ В EXCEL ============

@app.route('/export-journal/<int:site_id>')
@login_required
def export_journal(site_id):
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    work_type = request.args.get('work_type')
    responsible = request.args.get('responsible')
    
    query = JournalEntry.query.filter_by(site_id=site_id)
    
    if date_from:
        query = query.filter(JournalEntry.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(JournalEntry.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if work_type:
        query = query.filter_by(work_type=work_type)
    if responsible:
        query = query.filter_by(responsible_person=responsible)
    
    entries = query.order_by(JournalEntry.date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал работ"
    
    headers = ["Дата", "Вид работ", "Описание", "Место", "Исполнитель", 
               "Ответственный", "Кол-во рабочих", "Начало", "Окончание", 
               "Объём", "Ед. изм.", "Оборудование", "Материалы", 
               "Погода", "Температура", "Примечания"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1a3a3a", end_color="1a3a3a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for entry in entries:
        ws.append([
            entry.date.strftime("%d.%m.%Y") if entry.date else "",
            entry.work_type,
            entry.work_description,
            entry.location,
            entry.executor,
            entry.responsible_person,
            entry.workers_count,
            entry.start_time,
            entry.end_time,
            entry.volume,
            entry.volume_unit,
            entry.equipment_used,
            entry.materials_used,
            entry.weather,
            entry.temperature,
            entry.notes
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    site_name = Site.query.get(site_id).name
    return send_file(output, as_attachment=True, download_name=f"Журнал_{site_name}_{datetime.now().strftime('%Y%m%d')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ============ ОТОБРАЖЕНИЕ ЗАГРУЖЕННЫХ ФАЙЛОВ ============

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ============ ЖУРНАЛ РАБОТ ============

@app.route('/journal/<int:site_id>')
@login_required
def journal(site_id):
    site = Site.query.get_or_404(site_id)
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    work_type = request.args.get('work_type')
    responsible = request.args.get('responsible')
    
    query = JournalEntry.query.filter_by(site_id=site_id)
    
    if date_from:
        query = query.filter(JournalEntry.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(JournalEntry.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if work_type:
        query = query.filter_by(work_type=work_type)
    if responsible:
        query = query.filter_by(responsible_person=responsible)
    
    page = request.args.get('page', 1, type=int)
    per_page = 15
    pagination = query.order_by(JournalEntry.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    entries = pagination.items
    
    total_entries = query.count()
    week_ago = datetime.now().date() - timedelta(days=7)
    weekly_count = JournalEntry.query.filter(
        JournalEntry.site_id == site_id,
        JournalEntry.date >= week_ago
    ).count()
    
    all_work_types = [wt[0] for wt in db.session.query(JournalEntry.work_type).filter_by(site_id=site_id).distinct()]
    responsible_persons = [r[0] for r in db.session.query(JournalEntry.responsible_person).filter_by(site_id=site_id).distinct() if r[0]]
    
    request_args = request.args.to_dict()
    if 'page' in request_args:
        del request_args['page']
    
    return render_template('journal.html',
                         site=site,
                         entries=entries,
                         pagination=pagination,
                         stats={'total_entries': total_entries, 'weekly_count': weekly_count},
                         all_work_types=all_work_types,
                         responsible_persons=responsible_persons,
                         date_from=date_from or '',
                         date_to=date_to or '',
                         selected_work_type=work_type or '',
                         selected_responsible=responsible or '',
                         request_args=request_args)

@app.route('/add-entry', methods=['POST'])
@login_required
def add_entry():
    data = request.form
    
    entry = JournalEntry(
        site_id=int(data['site_id']),
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        work_type=data['work_type'],
        work_description=data.get('work_description', ''),
        location=data.get('location', ''),
        executor=data.get('executor', ''),
        responsible_person=data.get('responsible_person', ''),
        workers_count=int(data.get('workers_count', 1)),
        start_time=data.get('start_time', ''),
        end_time=data.get('end_time', ''),
        volume=float(data.get('volume', 0)),
        volume_unit=data.get('volume_unit', 'м³'),
        equipment_used=data.get('equipment_used', ''),
        materials_used=data.get('materials_used', ''),
        weather=data.get('weather', ''),
        temperature=data.get('temperature', ''),
        notes=data.get('notes', '')
    )
    
    db.session.add(entry)
    db.session.commit()
    flash('✅ Запись успешно добавлена', 'success')
    return redirect(url_for('journal', site_id=data['site_id']))

@app.route('/get-entry/<int:entry_id>')
@login_required
def get_entry(entry_id):
    entry = JournalEntry.query.get_or_404(entry_id)
    return jsonify(entry.to_dict())

@app.route('/update-entry/<int:entry_id>', methods=['POST'])
@login_required
def update_entry(entry_id):
    entry = JournalEntry.query.get_or_404(entry_id)
    data = request.form
    
    entry.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    entry.work_type = data['work_type']
    entry.work_description = data.get('work_description', '')
    entry.location = data.get('location', '')
    entry.executor = data.get('executor', '')
    entry.responsible_person = data.get('responsible_person', '')
    entry.workers_count = int(data.get('workers_count', 1))
    entry.start_time = data.get('start_time', '')
    entry.end_time = data.get('end_time', '')
    entry.volume = float(data.get('volume', 0))
    entry.volume_unit = data.get('volume_unit', 'м³')
    entry.equipment_used = data.get('equipment_used', '')
    entry.materials_used = data.get('materials_used', '')
    entry.weather = data.get('weather', '')
    entry.temperature = data.get('temperature', '')
    entry.notes = data.get('notes', '')
    
    db.session.commit()
    flash('✏️ Запись успешно обновлена', 'success')
    return redirect(url_for('journal', site_id=entry.site_id))

@app.route('/delete-entry/<int:entry_id>')
@login_required
def delete_entry(entry_id):
    entry = JournalEntry.query.get_or_404(entry_id)
    site_id = entry.site_id
    db.session.delete(entry)
    db.session.commit()
    flash('🗑️ Запись удалена', 'danger')
    return redirect(url_for('journal', site_id=site_id))

# ============ НОРМАТИВЫ ============

@app.route('/normatives')
@login_required
def normatives():
    search = request.args.get('search', '')
    doc_type = request.args.get('type', '')
    page = request.args.get('page', 1, type=int)
    per_page = 12
    
    query = Normative.query
    
    if search:
        query = query.filter(
            db.or_(
                Normative.code.ilike(f'%{search}%'),
                Normative.title.ilike(f'%{search}%'),
                Normative.description.ilike(f'%{search}%'),
                Normative.tags.ilike(f'%{search}%')
            )
        )
    
    if doc_type:
        query = query.filter(Normative.doc_type == doc_type)
    
    pagination = query.order_by(Normative.is_favorite.desc(), Normative.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    normatives_list = pagination.items
    
    categories = db.session.query(Normative.doc_type).distinct().all()
    
    return render_template('normatives.html', 
                         normatives=normatives_list,
                         pagination=pagination,
                         search=search,
                         selected_type=doc_type,
                         categories=[c[0] for c in categories if c[0]])

@app.route('/add-normative', methods=['POST'])
@login_required
def add_normative():
    data = request.form
    normative = Normative(
        code=data.get('code', ''),
        title=data.get('title', ''),
        description=data.get('description', ''),
        doc_type=data.get('doc_type', ''),
        category=data.get('category', ''),
        status=data.get('status', 'действует'),
        actual_date=data.get('actual_date', ''),
        tags=data.get('tags', ''),
        file_url=data.get('file_url', '')
    )
    db.session.add(normative)
    db.session.commit()
    flash('✅ Нормативный документ добавлен', 'success')
    return redirect(url_for('normatives'))

@app.route('/normative/<int:doc_id>')
@login_required
def normative_view(doc_id):
    doc = Normative.query.get_or_404(doc_id)
    doc.views += 1
    db.session.commit()
    return render_template('normative_view.html', doc=doc)

@app.route('/favorite-toggle/<int:doc_id>')
@login_required
def favorite_toggle(doc_id):
    doc = Normative.query.get_or_404(doc_id)
    doc.is_favorite = not doc.is_favorite
    db.session.commit()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'is_favorite': doc.is_favorite})
    
    if doc.is_favorite:
        flash('⭐ Документ добавлен в избранное', 'success')
    else:
        flash('⭐ Документ удалён из избранного', 'info')
    return redirect(request.referrer or url_for('normatives'))

@app.route('/delete-normative/<int:doc_id>')
@login_required
@admin_required
def delete_normative(doc_id):
    doc = Normative.query.get_or_404(doc_id)
    db.session.delete(doc)
    db.session.commit()
    flash('🗑️ Документ удалён', 'danger')
    return redirect(url_for('normatives'))

# ============ БАЗА ЗНАНИЙ ============

@app.route('/knowledge_base')
@login_required
def knowledge_base():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    
    query = KnowledgeArticle.query
    
    if search:
        query = query.filter(
            db.or_(
                KnowledgeArticle.title.ilike(f'%{search}%'),
                KnowledgeArticle.content.ilike(f'%{search}%')
            )
        )
    
    if category:
        query = query.filter(KnowledgeArticle.category == category)
    
    articles = query.order_by(KnowledgeArticle.created_at.desc()).all()
    categories = db.session.query(KnowledgeArticle.category).distinct().all()
    
    return render_template('knowledge_base.html', 
                         articles=articles, 
                         categories=[c[0] for c in categories if c[0]],
                         search=search,
                         selected_category=category)

@app.route('/article/<int:article_id>')
@login_required
def view_article(article_id):
    article = KnowledgeArticle.query.get_or_404(article_id)
    article.views += 1
    db.session.commit()
    return render_template('article.html', article=article)

@app.route('/add-article', methods=['POST'])
@login_required
def add_article():
    data = request.form
    article = KnowledgeArticle(
        title=data['title'],
        content=data['content'],
        category=data.get('category', ''),
        author=session.get('username', 'Admin')
    )
    db.session.add(article)
    db.session.commit()
    flash('✅ Статья добавлена', 'success')
    return redirect(url_for('knowledge_base'))

@app.route('/delete-article/<int:article_id>')
@login_required
@admin_required
def delete_article(article_id):
    article = KnowledgeArticle.query.get_or_404(article_id)
    db.session.delete(article)
    db.session.commit()
    flash('🗑️ Статья удалена', 'danger')
    return redirect(url_for('knowledge_base'))
# ============ СТРАНИЦЫ РАЗДЕЛОВ ОБЪЕКТА ============

@app.route('/site/<int:site_id>/incoming')
@login_required
def site_incoming(site_id):
    site = Site.query.get_or_404(site_id)
    stats = {
        'incoming_count': Document.query.filter_by(site_id=site_id, doc_type='incoming').count()
    }
    return render_template('site_incoming.html', site=site, stats=stats)

@app.route('/site/<int:site_id>/outgoing')
@login_required
def site_outgoing(site_id):
    site = Site.query.get_or_404(site_id)
    stats = {
        'outgoing_count': Document.query.filter_by(site_id=site_id, doc_type='outgoing').count()
    }
    return render_template('site_outgoing.html', site=site, stats=stats)

@app.route('/site/<int:site_id>/executive')
@login_required
def site_executive(site_id):
    site = Site.query.get_or_404(site_id)
    stats = {
        'executive_count': Document.query.filter_by(site_id=site_id, doc_type='executive').count()
    }
    return render_template('site_executive.html', site=site, stats=stats)

@app.route('/site/<int:site_id>/media')
@login_required
def site_media(site_id):
    site = Site.query.get_or_404(site_id)
    return render_template('site_media.html', site=site)

@app.route('/site/<int:site_id>/reports')
@login_required
def site_reports(site_id):
    site = Site.query.get_or_404(site_id)
    return render_template('site_reports.html', site=site)

@app.route('/site/<int:site_id>/schedule')
@login_required
def site_schedule(site_id):
    site = Site.query.get_or_404(site_id)
    return render_template('site_schedule.html', site=site)

# ============ РОУТЫ КАЛЬКУЛЯТОРА ============

@app.route('/site/<int:site_id>/calculator')
@login_required
def site_calculator(site_id):
    """Страница калькулятора материалов для объекта"""
    site = Site.query.get_or_404(site_id)
    
    # Группируем материалы по категориям для удобного отображения
    categories = {}
    for key, material in CONSTRUCTION_MATERIALS.items():
        cat = material.get("category", "Прочие работы")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append({"key": key, "name": material["name"]})
    
    # Отдельно собираем металлопрокат
    metal_items = [{"key": k, "name": v["name"], "weight_per_m": v.get("weight_per_m"), "type": v.get("type", "profile")} 
                   for k, v in METAL_SECTIONS.items()]
    
    # Отдельно асфальт
    asphalt_items = [{"key": k, "name": v["name"], "consumption_per_cm": v["consumption_per_cm"]} 
                     for k, v in ASPHALT_TYPES.items()]
    
    # Получаем сохранённые расчёты из сессии
    saved_calculations = session.get(f'calculations_{site_id}', [])
    
    return render_template('site_calculator.html', 
                         site=site, 
                         categories=categories,
                         metal_items=metal_items,
                         asphalt_items=asphalt_items,
                         saved_calculations=saved_calculations,
                         construction_materials=CONSTRUCTION_MATERIALS)

@app.route('/api/calculate/construction', methods=['POST'])
@login_required
def api_calculate_construction():
    """API для расчёта общестроительных работ"""
    data = request.json
    material_key = data.get('material_key')
    quantity = float(data.get('quantity', 0))
    price_per_unit = float(data.get('price_per_unit', 0))
    
    if material_key not in CONSTRUCTION_MATERIALS:
        return jsonify({'error': 'Материал не найден'}), 400
    
    material = CONSTRUCTION_MATERIALS[material_key]
    
    # Основной материал
    main_qty_net = quantity * material["main_material"]["per_unit"]
    main_qty_with_waste = main_qty_net * material["waste_factor"]
    main_price = main_qty_with_waste * price_per_unit
    
    result = {
        "work_name": material["name"],
        "quantity": quantity,
        "unit": material["unit"],
        "main_material": {
            "name": material["main_material"]["name"],
            "qty_net": round(main_qty_net, 2),
            "qty_with_waste": round(main_qty_with_waste, 2),
            "unit": material["main_material"]["unit"],
            "price": round(main_price, 2)
        },
        "additional_materials": [],
        "total_price": round(main_price, 2)
    }
    
    # Дополнительные материалы
    if material.get("additional"):
        for add_mat in material["additional"]:
            add_qty_net = quantity * add_mat["per_unit"]
            add_qty_with_waste = add_qty_net * material["waste_factor"]
            # Для доп. материалов цену пока не считаем (можно будет добавить позже)
            result["additional_materials"].append({
                "name": add_mat["name"],
                "qty_net": round(add_qty_net, 2),
                "qty_with_waste": round(add_qty_with_waste, 2),
                "unit": add_mat["unit"]
            })
    
    return jsonify(result)

@app.route('/api/calculate/metal', methods=['POST'])
@login_required
def api_calculate_metal():
    """API для расчёта металлопроката"""
    data = request.json
    metal_key = data.get('metal_key')
    length = float(data.get('length', 0))  # в метрах
    price_per_kg = float(data.get('price_per_kg', 0))
    
    if metal_key not in METAL_SECTIONS:
        return jsonify({'error': 'Металлоизделие не найдено'}), 400
    
    metal = METAL_SECTIONS[metal_key]
    
    # Для листового металла - особый расчёт
    if metal.get('type') == 'sheet':
        thickness_cm = float(data.get('thickness', 0)) / 1000  # переводим мм в метры
        area = length  # для листа length = площадь в м²
        volume = area * thickness_cm
        weight = volume * metal['density']
        result = {
            "work_name": metal["name"],
            "length": area,
            "length_unit": "м²",
            "thickness": data.get('thickness', 0),
            "weight_kg": round(weight, 2),
            "price_total": round(weight * price_per_kg, 2)
        }
    else:
        # Для профильного проката
        weight = length * metal["weight_per_m"]
        result = {
            "work_name": metal["name"],
            "length": length,
            "length_unit": "м",
            "weight_kg": round(weight, 2),
            "price_total": round(weight * price_per_kg, 2),
            "weight_per_m": metal["weight_per_m"]
        }
    
    return jsonify(result)

@app.route('/api/calculate/asphalt', methods=['POST'])
@login_required
def api_calculate_asphalt():
    """API для расчёта асфальтирования"""
    data = request.json
    asphalt_key = data.get('asphalt_key')
    area = float(data.get('area', 0))  # м²
    thickness = float(data.get('thickness', 5))  # см
    price_per_kg = float(data.get('price_per_kg', 0))
    
    if asphalt_key not in ASPHALT_TYPES:
        return jsonify({'error': 'Тип асфальта не найден'}), 400
    
    asphalt = ASPHALT_TYPES[asphalt_key]
    
    # Расход на заданную толщину
    consumption_kg = area * asphalt["consumption_per_cm"] * thickness
    total_price = consumption_kg * price_per_kg / 1000  # цена за тонну, переводим в кг
    
    result = {
        "work_name": asphalt["name"],
        "area": area,
        "thickness": thickness,
        "consumption_kg": round(consumption_kg, 2),
        "consumption_tons": round(consumption_kg / 1000, 3),
        "price_total": round(total_price, 2)
    }
    
    return jsonify(result)

@app.route('/api/save_calculation/<int:site_id>', methods=['POST'])
@login_required
def save_calculation(site_id):
    """Сохраняет расчёт в сессию"""
    data = request.json
    
    saved_calculations = session.get(f'calculations_{site_id}', [])
    
    calculation = {
        "id": len(saved_calculations) + 1,
        "date": datetime.now().strftime('%d.%m.%Y %H:%M'),
        "work_name": data.get('work_name'),
        "quantity": data.get('quantity'),
        "unit": data.get('unit'),
        "total_quantity": data.get('total_quantity'),  # 👈 ДОБАВЛЕНО
        "unit_name": data.get('unit_name'),            # 👈 ДОБАВЛЕНО
        "total_price": data.get('total_price', 0),
        "details": data.get('details', {})
    }
    
    saved_calculations.insert(0, calculation)
    if len(saved_calculations) > 50:
        saved_calculations = saved_calculations[:50]
    
    session[f'calculations_{site_id}'] = saved_calculations
    session.modified = True
    
    return jsonify({'success': True, 'calculation': calculation})

@app.route('/api/clear_calculations/<int:site_id>', methods=['POST'])
@login_required
def clear_calculations(site_id):
    """Очищает историю расчётов в сессии"""
    session[f'calculations_{site_id}'] = []
    session.modified = True
    return jsonify({'success': True})

# ============ БЫСТРЫЙ КАЛЬКУЛЯТОР (ГЛОБАЛЬНЫЙ) ============

@app.route('/quick-calculator')
@login_required
def quick_calculator():
    """Быстрый калькулятор — для срочных расчётов без привязки к объекту"""
    # Группируем материалы по категориям
    categories = {}
    for key, material in CONSTRUCTION_MATERIALS.items():
        cat = material.get("category", "Прочие работы")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append({"key": key, "name": material["name"]})
    
    metal_items = [{"key": k, "name": v["name"], "weight_per_m": v.get("weight_per_m"), "type": v.get("type", "profile")} 
                   for k, v in METAL_SECTIONS.items()]
    
    asphalt_items = [{"key": k, "name": v["name"], "consumption_per_cm": v["consumption_per_cm"]} 
                     for k, v in ASPHALT_TYPES.items()]
    
    return render_template('quick_calculator.html',
                         categories=categories,
                         metal_items=metal_items,
                         asphalt_items=asphalt_items)

# ============ API ГРАФИКА РАБОТ ============

@app.route('/api/schedule/<int:site_id>', methods=['GET'])
@login_required
def get_schedule(site_id):
    """Получить все задачи графика"""
    tasks = ScheduleTask.query.filter_by(site_id=site_id).order_by(ScheduleTask.order_num).all()
    return jsonify([t.to_dict() for t in tasks])

@app.route('/api/schedule/add', methods=['POST'])
@login_required
def add_schedule_task():
    """Добавить задачу"""
    data = request.json
    task = ScheduleTask(
        site_id=data['site_id'],
        order_num=data.get('order_num', 0),
        task_name=data['task_name'],
        description=data.get('description', ''),
        start_date=datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None,
        end_date=datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None,
        duration=data.get('duration', 0),
        dependencies=data.get('dependencies', ''),
        status=data.get('status', 'plan'),
        progress=data.get('progress', 0),
        required_materials=data.get('required_materials', ''),
        responsible=data.get('responsible', ''),
        notes=data.get('notes', '')
    )
    db.session.add(task)
    db.session.commit()
    return jsonify({'success': True, 'id': task.id})

@app.route('/api/schedule/update/<int:task_id>', methods=['PUT'])
@login_required
def update_schedule_task(task_id):
    """Обновить задачу"""
    task = ScheduleTask.query.get_or_404(task_id)
    data = request.json
    
    task.task_name = data.get('task_name', task.task_name)
    task.description = data.get('description', task.description)
    task.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None
    task.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None
    task.duration = data.get('duration', task.duration)
    task.dependencies = data.get('dependencies', task.dependencies)
    task.status = data.get('status', task.status)
    task.progress = data.get('progress', task.progress)
    task.required_materials = data.get('required_materials', task.required_materials)
    task.responsible = data.get('responsible', task.responsible)
    task.notes = data.get('notes', task.notes)
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/schedule/delete/<int:task_id>', methods=['DELETE'])
@login_required
def delete_schedule_task(task_id):
    """Удалить задачу"""
    task = ScheduleTask.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/schedule/import/<int:site_id>', methods=['POST'])
@login_required
def import_schedule_excel(site_id):
    """Импорт графика из Excel"""
    from openpyxl import load_workbook
    
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не выбран'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    
    ScheduleTask.query.filter_by(site_id=site_id).delete()
    
    wb = load_workbook(file)
    ws = wb.active
    
    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row[0]:
            continue
        
        task = ScheduleTask(
            site_id=site_id,
            order_num=row_idx,
            task_name=str(row[0]) if row[0] else '',
            description=str(row[1]) if row[1] else '',
            start_date=row[2] if row[2] else None,
            end_date=row[3] if row[3] else None,
            duration=int(row[4]) if row[4] else 0,
            dependencies=str(row[5]) if row[5] else '',
            status=str(row[6]) if row[6] else 'plan',
            progress=int(row[7]) if row[7] else 0,
            responsible=str(row[8]) if row[8] else '',
            notes=str(row[9]) if row[9] else ''
        )
        tasks.append(task)
    
    db.session.add_all(tasks)
    db.session.commit()
    
    return jsonify({'success': True, 'count': len(tasks)})

@app.route('/api/schedule/analyze/<int:site_id>')
@login_required
def analyze_schedule(site_id):
    """Анализ рисков и цепных реакций"""
    tasks = ScheduleTask.query.filter_by(site_id=site_id).order_by(ScheduleTask.order_num).all()
    today = datetime.now().date()
    
    risks = []
    warnings = []
    notifications = []
    
    for task in tasks:
        # Отставание
        if task.end_date and task.end_date < today and task.progress < 100:
            delay_days = (today - task.end_date).days
            risks.append({
                'level': 'danger',
                'task_id': task.id,
                'task_name': task.task_name,
                'message': f'🔴 Отставание на {delay_days} дн.! План: {task.end_date}, Прогресс: {task.progress}%'
            })
        
        # Скорый старт
        if task.start_date and task.status == 'plan':
            days_to_start = (task.start_date - today).days
            if 0 <= days_to_start <= 3:
                notifications.append({
                    'level': 'warning',
                    'task_id': task.id,
                    'task_name': task.task_name,
                    'message': f'🟡 Стартует через {days_to_start} дн.! (с {task.start_date})'
                })
        
        # Зависимости
        if task.dependencies and task.status != 'completed':
            dep_ids = [int(x.strip()) for x in task.dependencies.split(',') if x.strip()]
            for dep_id in dep_ids:
                dep_task = ScheduleTask.query.get(dep_id)
                if dep_task and dep_task.end_date and dep_task.progress < 100:
                    if dep_task.end_date > (task.start_date or today):
                        warnings.append({
                            'level': 'warning',
                            'task_id': task.id,
                            'task_name': task.task_name,
                            'message': f'🟠 Зависит от "{dep_task.task_name}" (задерживается)'
                        })
        
        # Прогресс
        if task.status == 'active' and task.progress >= 50 and task.progress < 100 and task.end_date:
            days_left = (task.end_date - today).days
            if days_left <= 5:
                warnings.append({
                    'level': 'info',
                    'task_id': task.id,
                    'task_name': task.task_name,
                    'message': f'📊 Выполнено на {task.progress}%, осталось {days_left} дн.'
                })
        
        # Материалы
        if task.required_materials and task.start_date:
            days_to_start = (task.start_date - today).days
            if 3 <= days_to_start <= 7:
                notifications.append({
                    'level': 'info',
                    'task_id': task.id,
                    'task_name': task.task_name,
                    'message': f'📦 Закажите материалы: {task.required_materials[:50]}'
                })
    
    total_progress = sum(t.progress for t in tasks) // len(tasks) if tasks else 0
    
    return jsonify({
        'risks': risks,
        'notifications': notifications,
        'warnings': warnings,
        'summary': {
            'total_tasks': len(tasks),
            'completed_tasks': sum(1 for t in tasks if t.progress >= 100),
            'active_tasks': sum(1 for t in tasks if t.status == 'active'),
            'delayed_tasks': len(risks),
            'overall_progress': total_progress
        },
        'last_update': datetime.now().strftime('%d.%m.%Y %H:%M')
    })

@app.route('/api/schedule/export/<int:site_id>')
@login_required
def export_schedule_excel(site_id):
    """Экспорт графика в Excel"""
    from openpyxl.utils import get_column_letter
    
    tasks = ScheduleTask.query.filter_by(site_id=site_id).order_by(ScheduleTask.order_num).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "График работ"
    
    headers = ["№", "Наименование работы", "Описание", "Начало", "Окончание", 
               "Длит.", "Зависимости", "Статус", "Прогресс %", "Материалы", "Ответственный", "Примечания"]
    ws.append(headers)
    
    for task in tasks:
        ws.append([
            task.order_num,
            task.task_name,
            task.description or '',
            task.start_date.strftime('%d.%m.%Y') if task.start_date else '',
            task.end_date.strftime('%d.%m.%Y') if task.end_date else '',
            task.duration or '',
            task.dependencies or '',
            task.status,
            task.progress,
            task.required_materials or '',
            task.responsible or '',
            task.notes or ''
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    site = Site.query.get(site_id)
    return send_file(output, as_attachment=True, 
                     download_name=f"График_работ_{site.name}_{datetime.now().strftime('%Y%m%d')}.xlsx")

if __name__ == '__main__':
    print("=" * 50)
    print("🚀 СтройДок - Строительный портал")
    print("=" * 50)
    print("📁 База данных:", app.config['SQLALCHEMY_DATABASE_URI'])
    print("🔑 Тестовый вход: admin / admin123")
    print("=" * 50)
    app.run(debug=True)